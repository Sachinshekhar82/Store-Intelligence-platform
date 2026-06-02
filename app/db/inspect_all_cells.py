import openpyxl

def run():
    wb = openpyxl.load_workbook("data/store_layout.xlsx", data_only=True)
    for name in wb.sheetnames:
        sheet = wb[name]
        print(f"Sheet: {name}, dimensions: {sheet.dimensions}")
        for r in range(1, sheet.max_row + 1):
            row_vals = []
            for c in range(1, sheet.max_column + 1):
                val = sheet.cell(row=r, column=c).value
                if val is not None:
                    row_vals.append(f"C{c}: {val}")
            if row_vals:
                print(f"Row {r:02d}: " + " | ".join(row_vals))

if __name__ == "__main__":
    run()
